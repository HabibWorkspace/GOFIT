"""Complete admin routes for member and trainer management."""
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required
from models.user import User, UserRole
from models.member_profile import MemberProfile
from models.trainer_profile import TrainerProfile
from models.package import Package
from models.transaction import Transaction, TransactionStatus, TransactionType
from services.password_service import PasswordService
from middleware.rbac import require_admin
from database import db
from utils.audit import log_action, get_change_details
from datetime import datetime, timedelta
from sqlalchemy import func
import uuid
from io import BytesIO

admin_complete_bp = Blueprint('admin_complete', __name__)

# ============================================================================
# MEMBER ROUTES
# ============================================================================

@admin_complete_bp.route('/members', methods=['GET'])
@require_admin
def list_members():
    """List all members with pagination and optional search."""
    from datetime import datetime, timedelta
    from models.attendance import Attendance
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    
    query = MemberProfile.query
    
    # Apply search filter if provided
    if search:
        search_lower = f'{search.lower()}%'  # starts-with only
        # For numeric search (member ID), use exact match first
        try:
            search_int = int(search)
            query = query.filter(
                db.or_(
                    MemberProfile.member_number == search_int,  # exact ID match
                    db.func.lower(MemberProfile.full_name).like(search_lower),
                    MemberProfile.phone.like(f'{search}%')
                )
            )
        except ValueError:
            # Non-numeric: search name and phone with starts-with
            query = query.filter(
                db.or_(
                    db.func.lower(MemberProfile.full_name).like(search_lower),
                    MemberProfile.phone.like(f'{search}%')
                )
            )
    
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Calculate 10 days ago for inactive check
    now = datetime.utcnow()
    ten_days_ago = now - timedelta(days=10)
    
    members = []
    for member in paginated.items:
        member_dict = member.to_dict()
        user = User.query.get(member.user_id)
        if user:
            member_dict['username'] = user.username
        
        # Add trainer information if assigned
        if member.trainer_id:
            trainer = TrainerProfile.query.get(member.trainer_id)
            if trainer:
                member_dict['trainer_name'] = trainer.full_name
                member_dict['trainer_specialization'] = trainer.specialization
            else:
                member_dict['trainer_name'] = None
        else:
            member_dict['trainer_name'] = None
        
        # Calculate inactive status
        # Member is inactive if: has attended before BUT not in last 10 days
        is_inactive = False
        if not member.is_frozen:  # Only check for active members
            # Check if member has ANY attendance record
            has_attended_before = Attendance.query.filter(
                Attendance.member_id == member.id
            ).first()
            
            if has_attended_before:
                # Check if member has recent attendance (last 10 days)
                recent_attendance = Attendance.query.filter(
                    Attendance.member_id == member.id,
                    Attendance.check_in_time >= ten_days_ago
                ).first()
                
                # If attended before but not recently, mark as inactive
                if not recent_attendance:
                    is_inactive = True
        
        member_dict['is_inactive'] = is_inactive
        members.append(member_dict)
    
    return jsonify({
        'members': members,
        'total': paginated.total,
        'page': page,
        'per_page': per_page
    }), 200


@admin_complete_bp.route('/members/export', methods=['GET'])
@require_admin
def export_members_excel():
    """Export all members to Excel file with GOFIT theme - Fully optimized."""
    try:
        # Lazy import openpyxl only when needed
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from sqlalchemy.orm import joinedload
        
        # FULLY OPTIMIZED: Use joinedload to eagerly fetch related data in a single query
        # This eliminates ALL N+1 queries by using SQL JOINs
        all_members = MemberProfile.query.options(
            joinedload(MemberProfile.trainer)
        ).order_by(MemberProfile.member_number).all()
        
        # Fetch all packages in one query (can't use joinedload since it's not a direct relationship)
        package_ids = [m.current_package_id for m in all_members if m.current_package_id]
        packages_dict = {}
        if package_ids:
            packages = Package.query.filter(Package.id.in_(package_ids)).all()
            packages_dict = {p.id: p.name for p in packages}
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Members List"
        
        # Define headers
        headers = [
            'Member ID', 'Full Name', 'Phone', 'Gender', 
            'Date of Birth', 'Admission Date', 'Current Package', 'Trainer'
        ]
        
        # OPTIMIZED: Create style objects once and reuse them
        header_fill = PatternFill(start_color='F2C228', end_color='F2C228', fill_type='solid')
        header_font = Font(bold=True, size=12, color='1A1A1A')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Reusable styles for data cells
        data_alignment = Alignment(horizontal='left', vertical='center')
        data_font = Font(size=11, color='1A1A1A')
        border = Border(
            left=Side(style='thin', color='1A1A1A'),
            right=Side(style='thin', color='1A1A1A'),
            top=Side(style='thin', color='1A1A1A'),
            bottom=Side(style='thin', color='1A1A1A')
        )
        row_fill_even = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Add title row
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'GOFIT - MEMBERS LIST'
        title_cell.font = Font(bold=True, size=16, color='1A1A1A')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = header_fill
        ws.row_dimensions[1].height = 30
        
        # Add headers in row 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        ws.row_dimensions[2].height = 25
        
        # Prepare all data first (no database queries in loop)
        rows_data = []
        for member in all_members:
            # Get package name from pre-fetched dictionary (no query)
            package_name = packages_dict.get(member.current_package_id, 'Not Assigned')
            
            # Get trainer name from eagerly loaded relationship (no query)
            trainer_name = member.trainer.full_name if member.trainer else 'Not Assigned'
            
            # Format dates
            dob = member.date_of_birth.strftime('%d-%b-%Y') if member.date_of_birth else 'N/A'
            admission = member.admission_date.strftime('%d-%b-%Y') if member.admission_date else 'N/A'
            
            rows_data.append([
                member.member_number or 'N/A',
                member.full_name or 'N/A',
                member.phone or 'N/A',
                member.gender or 'N/A',
                dob,
                admission,
                package_name,
                trainer_name
            ])
        
        # OPTIMIZED: Batch write all rows at once
        for row_num, row_data in enumerate(rows_data, 3):
            row_fill = row_fill_even if row_num % 2 == 0 else row_fill_odd
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = data_alignment
                cell.border = border
                cell.fill = row_fill
                cell.font = data_font
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log the action
        log_action(
            action='exported members to Excel',
            target_type='Export',
            target_id=None,
            details={
                'total_members': len(all_members),
                'filename': f"GOFIT_Members_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
        # Generate filename
        filename = f"GOFIT_Members_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        current_app.logger.error(f"Error exporting members to Excel: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/export', methods=['POST'])
@require_admin
def export_finance_excel():
    """Export finance transactions to Excel file."""
    try:
        # Lazy import openpyxl only when needed
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        data = request.get_json()
        transactions = data.get('transactions', [])
        selected_month = data.get('month', '')
        
        if not transactions:
            return jsonify({'error': 'No transactions to export'}), 400
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Finance Report"
        
        # Define headers - simplified (removed Phone, Status, Type)
        headers = [
            'Member ID', 'Member Name', 'Amount (Rs.)', 
            'Due Date', 'Paid Date'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        header_font = Font(bold=True, size=12, color='000000')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Add title row
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        month_text = ""
        if selected_month:
            try:
                month_date = datetime.strptime(selected_month + '-01', '%Y-%m-%d')
                month_text = f" - {month_date.strftime('%B %Y')}"
            except:
                pass
        title_cell.value = f'MODERN FITNESS GYM - FINANCE REPORT{month_text}'
        title_cell.font = Font(bold=True, size=16, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        ws.row_dimensions[1].height = 30
        
        # Add headers in row 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        ws.row_dimensions[2].height = 25
        
        # Add data rows starting from row 3
        total_amount = 0
        paid_count = 0
        
        for row_num, txn in enumerate(transactions, 3):
            # Format dates
            due_date = datetime.fromisoformat(txn['due_date'].replace('Z', '')).strftime('%d-%b-%Y') if txn.get('due_date') else 'N/A'
            paid_date = datetime.fromisoformat(txn['paid_date'].replace('Z', '')).strftime('%d-%b-%Y') if txn.get('paid_date') else 'N/A'
            
            # Write row data
            row_data = [
                txn.get('member_id', 'N/A'),
                txn.get('member_name', 'N/A'),
                float(txn.get('amount', 0)),
                due_date,
                paid_date
            ]
            
            # Alternate row colors
            row_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid') if row_num % 2 == 0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            # Highlight paid rows (has paid_date)
            if paid_date != 'N/A':
                row_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                total_amount += float(txn.get('amount', 0))
                paid_count += 1
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='left' if col_num != 3 else 'right', vertical='center')
                cell.border = border
                cell.fill = row_fill
                cell.font = Font(size=11)
                
                # Format amount as currency
                if col_num == 3:
                    cell.number_format = '#,##0.00'
        
        # Add summary rows
        summary_row = len(transactions) + 3
        
        # Total Collected row
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        total_label_cell = ws[f'A{summary_row}']
        total_label_cell.value = 'TOTAL COLLECTED'
        total_label_cell.font = Font(bold=True, size=12, color='000000')
        total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_label_cell.fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        total_label_cell.border = border
        
        total_amount_cell = ws[f'C{summary_row}']
        total_amount_cell.value = total_amount
        total_amount_cell.font = Font(bold=True, size=12, color='000000')
        total_amount_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_amount_cell.fill = PatternFill(start_color='B6FF00', end_color='B6FF00', fill_type='solid')
        total_amount_cell.border = border
        total_amount_cell.number_format = '#,##0.00'
        
        # Paid Count row
        summary_row += 1
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        count_label_cell = ws[f'A{summary_row}']
        count_label_cell.value = 'PAYMENTS RECEIVED'
        count_label_cell.font = Font(bold=True, size=11, color='000000')
        count_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        count_label_cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        count_label_cell.border = border
        
        count_cell = ws[f'C{summary_row}']
        count_cell.value = paid_count
        count_cell.font = Font(bold=True, size=11, color='000000')
        count_cell.alignment = Alignment(horizontal='right', vertical='center')
        count_cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        count_cell.border = border
        
        # Set column widths
        column_widths = {
            'A': 12,  # Member ID
            'B': 30,  # Member Name
            'C': 18,  # Amount
            'D': 18,  # Due Date
            'E': 18   # Paid Date
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        month_suffix = f"_{selected_month}" if selected_month else ""
        filename = f"Finance_Report{month_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/members/test-details', methods=['GET'])
def test_details_route():
    """Test route without variable."""
    return jsonify({'message': 'Test route works!'}), 200


@admin_complete_bp.route('/members/details/<member_id>', methods=['GET'])
def get_member_details(member_id):
    """Get comprehensive member details including history, transactions, and attendance."""
    # TEMPORARY DEBUG - NO AUTH
    return jsonify({'debug': 'Route is working!', 'member_id': member_id}), 200
    
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    # Basic member info
    member_dict = member.to_dict()
    user = User.query.get(member.user_id)
    if user:
        member_dict['username'] = user.username
    
    # Current package info
    if member.current_package_id:
        package = Package.query.get(member.current_package_id)
        if package:
            member_dict['current_package'] = {
                'id': package.id,
                'name': package.name,
                'price': float(package.price),
                'duration_days': package.duration_days,
                'description': package.description
            }
    
    # Current trainer info
    if member.trainer_id:
        trainer = TrainerProfile.query.get(member.trainer_id)
        if trainer:
            member_dict['current_trainer'] = {
                'id': trainer.id,
                'full_name': trainer.full_name,
                'specialization': trainer.specialization,
                'phone': trainer.phone,
                'email': trainer.email
            }
    
    # Get all transactions (payment history and membership periods)
    transactions = Transaction.query.filter_by(member_id=member_id).order_by(Transaction.due_date.desc()).all()
    member_dict['transactions'] = []
    
    for txn in transactions:
        txn_dict = txn.to_dict()
        
        # Add package info for each transaction
        if member.current_package_id:
            package = Package.query.get(member.current_package_id)
            if package:
                txn_dict['package_name'] = package.name
                txn_dict['package_duration'] = package.duration_days
        
        # Add trainer info if transaction has trainer fee
        if txn.trainer_fee and txn.trainer_fee > 0:
            if member.trainer_id:
                trainer = TrainerProfile.query.get(member.trainer_id)
                if trainer:
                    txn_dict['trainer_name'] = trainer.full_name
        
        member_dict['transactions'].append(txn_dict)
    
    # Build membership history from transactions
    # Group by package periods (when member had active membership)
    membership_history = []
    for txn in transactions:
        if txn.transaction_type == TransactionType.PACKAGE and txn.status == TransactionStatus.COMPLETED:
            # Calculate period based on due_date and package duration
            if member.current_package_id:
                package = Package.query.get(member.current_package_id)
                if package and txn.due_date:
                    start_date = txn.paid_date if txn.paid_date else txn.due_date
                    end_date = start_date + timedelta(days=package.duration_days)
                    
                    period = {
                        'package_name': package.name,
                        'start_date': start_date.isoformat() + 'Z',
                        'end_date': end_date.isoformat() + 'Z',
                        'amount_paid': float(txn.amount),
                        'payment_date': txn.paid_date.isoformat() + 'Z' if txn.paid_date else None,
                        'trainer_name': None
                    }
                    
                    # Add trainer info if available
                    if member.trainer_id:
                        trainer = TrainerProfile.query.get(member.trainer_id)
                        if trainer:
                            period['trainer_name'] = trainer.full_name
                    
                    membership_history.append(period)
    
    member_dict['membership_history'] = membership_history
    
    # Attendance statistics
    attendance_records = Attendance.query.filter_by(member_id=member_id).all()
    
    # Total check-ins
    total_checkins = len(attendance_records)
    
    # Monthly breakdown (last 12 months)
    monthly_attendance = {}
    current_date = datetime.utcnow()
    
    for i in range(12):
        month_date = current_date - timedelta(days=30 * i)
        month_key = month_date.strftime('%Y-%m')
        month_name = month_date.strftime('%b %Y')
        
        # Count attendance for this month
        month_count = Attendance.query.filter(
            Attendance.member_id == member_id,
            extract('year', Attendance.check_in_time) == month_date.year,
            extract('month', Attendance.check_in_time) == month_date.month
        ).count()
        
        monthly_attendance[month_key] = {
            'month': month_name,
            'count': month_count
        }
    
    # Calculate attendance rate (if member has active package)
    attendance_rate = 0
    days_attended = 0
    if member.package_start_date and member.package_expiry_date:
        # Days in current package period
        days_in_period = (member.package_expiry_date - member.package_start_date).days
        
        # Days attended in current period
        days_attended = Attendance.query.filter(
            Attendance.member_id == member_id,
            Attendance.check_in_time >= member.package_start_date,
            Attendance.check_in_time <= member.package_expiry_date
        ).distinct(func.date(Attendance.check_in_time)).count()
        
        if days_in_period > 0:
            attendance_rate = round((days_attended / days_in_period) * 100, 1)
    
    member_dict['attendance_stats'] = {
        'total_checkins': total_checkins,
        'monthly_breakdown': monthly_attendance,
        'attendance_rate': attendance_rate,
        'days_attended_current_period': days_attended
    }
    
    return jsonify(member_dict), 200


@admin_complete_bp.route('/members/<member_id>', methods=['GET'])
@require_admin
def get_member(member_id):
    """Get a specific member by ID."""
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    member_dict = member.to_dict()
    user = User.query.get(member.user_id)
    if user:
        member_dict['username'] = user.username
    
    # Add trainer information if assigned
    if member.trainer_id:
        trainer = TrainerProfile.query.get(member.trainer_id)
        if trainer:
            member_dict['trainer_name'] = trainer.full_name
            member_dict['trainer_specialization'] = trainer.specialization
        else:
            member_dict['trainer_name'] = None
    else:
        member_dict['trainer_name'] = None
    
    return jsonify(member_dict), 200


@admin_complete_bp.route('/members', methods=['POST'])
@require_admin
def create_member():
    """Create a new member."""
    data = request.get_json()
    
    # Validate required fields (only full_name and phone are required)
    required_fields = ['full_name', 'phone']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    try:
        # Create user with auto-generated username
        # Use member's name as username (lowercase, no spaces, alphanumeric only)
        import re
        auto_username = data['full_name'].lower().replace(' ', '')
        # Remove all non-alphanumeric characters except underscore
        auto_username = re.sub(r'[^a-z0-9_]', '', auto_username)
        # Ensure username is not empty and has minimum length
        if not auto_username or len(auto_username) < 3:
            # Fallback to member + random string if name is too short
            import uuid
            auto_username = f"member{str(uuid.uuid4())[:8]}"
        
        auto_password = 'member123'  # Default password for all members
        
        user = User(
            username=auto_username,
            password_hash=PasswordService.hash_password(auto_password),
            password=auto_password,  # Store plain password for display
            role=UserRole.MEMBER,
            is_active=True
        )
        db.session.add(user)
        db.session.flush()
        
        # Parse date_of_birth if provided
        dob = None
        if data.get('date_of_birth'):
            try:
                from datetime import datetime as dt
                dob = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            except:
                dob = None
        
        # Parse admission_date if provided
        admission_date = datetime.utcnow()
        if data.get('admission_date'):
            try:
                from datetime import datetime as dt
                admission_date = dt.strptime(data['admission_date'], '%Y-%m-%d')
            except:
                admission_date = datetime.utcnow()
        
        # Calculate package dates if package is assigned
        # Use provided dates if available, otherwise auto-calculate
        package_start_date = None
        package_expiry_date = None
        if data.get('package_id'):
            # Check if manual dates are provided
            if data.get('package_start_date'):
                try:
                    from datetime import datetime as dt
                    package_start_date = dt.strptime(data['package_start_date'], '%Y-%m-%d')
                except:
                    package_start_date = datetime.utcnow()
            else:
                package_start_date = datetime.utcnow()
            
            if data.get('package_expiry_date'):
                try:
                    from datetime import datetime as dt
                    package_expiry_date = dt.strptime(data['package_expiry_date'], '%Y-%m-%d')
                except:
                    # Auto-calculate if manual date fails
                    package = Package.query.get(data['package_id'])
                    if package:
                        package_expiry_date = package_start_date + timedelta(days=package.duration_days)
            else:
                # Auto-calculate if no manual date provided
                package = Package.query.get(data['package_id'])
                if package:
                    package_expiry_date = package_start_date + timedelta(days=package.duration_days)
        
        # Auto-generate member_number starting from 10
        # Get the highest member_number and increment by 1
        max_member_number = db.session.query(func.max(MemberProfile.member_number)).scalar()
        if max_member_number is None or max_member_number < 10:
            new_member_number = 10
        else:
            new_member_number = max_member_number + 1
        
        # Create member profile
        # Convert empty strings to None for optional fields
        cnic_value = data.get('cnic') if data.get('cnic') else None
        card_id_value = data.get('card_id') if data.get('card_id') else None
        email_value = data.get('email') if data.get('email') else None
        father_name_value = data.get('father_name') if data.get('father_name') else None
        blood_group_value = data.get('blood_group') if data.get('blood_group') else None
        address_value = data.get('address') if data.get('address') else None
        
        # Parse weight_kg if provided
        weight_kg_value = None
        if data.get('weight_kg'):
            try:
                weight_kg_value = float(data['weight_kg'])
            except:
                weight_kg_value = None
        
        member = MemberProfile(
            user_id=user.id,
            member_number=new_member_number,
            card_id=card_id_value,  # RFID card number
            full_name=data['full_name'],
            phone=data['phone'],
            cnic=cnic_value,  # None if empty
            email=email_value,
            father_name=father_name_value,
            weight_kg=weight_kg_value,
            blood_group=blood_group_value,
            address=address_value,
            gender=data.get('gender'),
            date_of_birth=dob,
            admission_date=admission_date,
            current_package_id=data.get('package_id') if data.get('package_id') else None,
            trainer_id=data.get('trainer_id') if data.get('trainer_id') else None,
            package_start_date=package_start_date,
            package_expiry_date=package_expiry_date,
            is_frozen=False,
            profile_picture=data.get('profile_picture')
        )
        db.session.add(member)
        db.session.flush()
        
        # Create admission fee transaction (PENDING status)
        # Due date should be the package expiry date if package is assigned, otherwise 7 days from now
        admission_fee = float(data.get('admission_fee', 5000))  # Use provided fee or default to 5000
        discount = float(data.get('discount', 0))
        discount_type = data.get('discount_type', 'fixed')  # 'fixed' or 'percentage'
        trainer_charge = float(data.get('trainer_charge', 0))
        
        # Get package price if package is assigned
        package_price = 0
        if data.get('package_id'):
            package = Package.query.get(data['package_id'])
            if package:
                package_price = float(package.price)
        
        # Calculate total amount for transaction
        # Total = admission_fee + package_price + trainer_charge
        total_amount = admission_fee + package_price + trainer_charge
        
        # Apply discount if any
        if discount > 0:
            if discount_type == 'percentage':
                discount_amount = total_amount * (discount / 100)
            else:
                discount_amount = discount
        else:
            discount_amount = 0
        
        # Final amount after discount
        final_amount = total_amount - discount_amount
        
        due_date = package_expiry_date if package_expiry_date else (datetime.utcnow() + timedelta(days=7))
        
        admission_transaction = Transaction(
            member_id=member.id,
            amount=final_amount,
            transaction_type=TransactionType.ADMISSION,
            status=TransactionStatus.PENDING,
            due_date=due_date,
            trainer_fee=trainer_charge,
            package_price=package_price,
            discount_amount=discount_amount,
            discount_type=discount_type,
            created_at=datetime.utcnow()
        )
        db.session.add(admission_transaction)
        
        # Log the action
        log_action(
            action='created member',
            target_type='Member',
            target_id=member.id,
            details={
                'member_number': new_member_number,
                'full_name': data['full_name'],
                'phone': data['phone'],
                'card_id': card_id_value,
                'package_id': data.get('package_id'),
                'trainer_id': data.get('trainer_id'),
                'admission_fee': final_amount,
                'discount': discount_amount
            }
        )
        
        db.session.commit()
        
        # Send welcome email if member has email address
        if email_value:
            try:
                from services.email_service import EmailService
                email_service = EmailService()
                email_service.send_welcome_email(
                    to_email=email_value,
                    member_name=data['full_name'],
                    username=auto_username,
                    password=auto_password
                )
                current_app.logger.info(f"Welcome email sent to {email_value}")
            except Exception as email_error:
                # Log error but don't fail the member creation
                current_app.logger.error(f"Failed to send welcome email: {str(email_error)}")
        
        member_dict = member.to_dict()
        
        return jsonify(member_dict), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/members/<member_id>', methods=['PUT'])
@require_admin
def update_member(member_id):
    """Update a member."""
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    data = request.get_json()
    
    try:
        # Capture state before changes for audit log
        before_state = {
            'full_name': member.full_name,
            'phone': member.phone,
            'cnic': member.cnic,
            'email': member.email,
            'father_name': member.father_name,
            'weight_kg': float(member.weight_kg) if member.weight_kg else None,
            'blood_group': member.blood_group,
            'address': member.address,
            'card_id': member.card_id,
            'current_package_id': member.current_package_id,
            'trainer_id': member.trainer_id,
            'package_start_date': member.package_start_date.isoformat() if member.package_start_date else None,
            'package_expiry_date': member.package_expiry_date.isoformat() if member.package_expiry_date else None,
            'is_frozen': member.is_frozen,
            'gender': member.gender,
            'date_of_birth': member.date_of_birth.isoformat() if member.date_of_birth else None,
            'admission_date': member.admission_date.isoformat() if member.admission_date else None
        }
        
        # Update member profile fields
        if 'full_name' in data:
            member.full_name = data['full_name']
        if 'phone' in data:
            member.phone = data['phone']
        if 'cnic' in data:
            member.cnic = data['cnic']
        if 'email' in data:
            member.email = data['email'] if data['email'] else None
        if 'father_name' in data:
            member.father_name = data['father_name'] if data['father_name'] else None
        if 'weight_kg' in data:
            try:
                member.weight_kg = float(data['weight_kg']) if data['weight_kg'] else None
            except:
                pass
        if 'blood_group' in data:
            member.blood_group = data['blood_group'] if data['blood_group'] else None
        if 'address' in data:
            member.address = data['address'] if data['address'] else None
        if 'card_id' in data:
            member.card_id = data['card_id']
        if 'is_frozen' in data:
            member.is_frozen = data['is_frozen']
        if 'profile_picture' in data:
            member.profile_picture = data['profile_picture']
        
        # Handle package assignment (accept both package_id and current_package_id)
        if 'package_id' in data:
            new_package_id = data['package_id'] if data['package_id'] else None
            # If package changed, recalculate dates
            if new_package_id != member.current_package_id and new_package_id:
                package = Package.query.get(new_package_id)
                if package:
                    # Only auto-set dates if not provided
                    if 'package_start_date' not in data or not data['package_start_date']:
                        member.package_start_date = datetime.utcnow()
                    if 'package_expiry_date' not in data or not data['package_expiry_date']:
                        start_date = member.package_start_date or datetime.utcnow()
                        member.package_expiry_date = start_date + timedelta(days=package.duration_days)
            member.current_package_id = new_package_id
        elif 'current_package_id' in data:
            new_package_id = data['current_package_id'] if data['current_package_id'] else None
            # If package changed, recalculate dates
            if new_package_id != member.current_package_id and new_package_id:
                package = Package.query.get(new_package_id)
                if package:
                    # Only auto-set dates if not provided
                    if 'package_start_date' not in data or not data['package_start_date']:
                        member.package_start_date = datetime.utcnow()
                    if 'package_expiry_date' not in data or not data['package_expiry_date']:
                        start_date = member.package_start_date or datetime.utcnow()
                        member.package_expiry_date = start_date + timedelta(days=package.duration_days)
            member.current_package_id = new_package_id
        
        # Handle package dates
        if 'package_start_date' in data:
            try:
                from datetime import datetime as dt
                member.package_start_date = dt.fromisoformat(data['package_start_date'].replace('Z', '+00:00')) if data['package_start_date'] else None
            except:
                pass
        
        if 'package_expiry_date' in data:
            try:
                from datetime import datetime as dt
                member.package_expiry_date = dt.fromisoformat(data['package_expiry_date'].replace('Z', '+00:00')) if data['package_expiry_date'] else None
            except:
                pass
        
        # Handle trainer assignment
        if 'trainer_id' in data:
            member.trainer_id = data['trainer_id'] if data['trainer_id'] else None
        
        if 'gender' in data:
            member.gender = data['gender']
        if 'date_of_birth' in data:
            try:
                from datetime import datetime as dt
                member.date_of_birth = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date() if data['date_of_birth'] else None
            except:
                pass
        
        # Handle admission_date
        if 'admission_date' in data:
            try:
                from datetime import datetime as dt
                member.admission_date = dt.strptime(data['admission_date'], '%Y-%m-%d').date() if data['admission_date'] else None
            except:
                pass
        
        # Capture state after changes for audit log
        after_state = {
            'full_name': member.full_name,
            'phone': member.phone,
            'cnic': member.cnic,
            'email': member.email,
            'father_name': member.father_name,
            'weight_kg': float(member.weight_kg) if member.weight_kg else None,
            'blood_group': member.blood_group,
            'address': member.address,
            'card_id': member.card_id,
            'current_package_id': member.current_package_id,
            'trainer_id': member.trainer_id,
            'package_start_date': member.package_start_date.isoformat() if member.package_start_date else None,
            'package_expiry_date': member.package_expiry_date.isoformat() if member.package_expiry_date else None,
            'is_frozen': member.is_frozen,
            'gender': member.gender,
            'date_of_birth': member.date_of_birth.isoformat() if member.date_of_birth else None,
            'admission_date': member.admission_date.isoformat() if member.admission_date else None
        }
        
        # Log the action with changes
        changes = get_change_details(before_state, after_state)
        if changes:
            log_action(
                action='updated member',
                target_type='Member',
                target_id=member.id,
                details={
                    'member_number': member.member_number,
                    'changes': changes
                }
            )
        
        # ── Auto-create PENDING transaction when package is newly assigned ──────
        # Only create if package changed AND package_expiry_date is set
        old_package_id = before_state.get('current_package_id')
        new_package_id = member.current_package_id
        package_changed = new_package_id and new_package_id != old_package_id

        if package_changed and member.package_expiry_date:
            package = Package.query.get(new_package_id)
            if package:
                # Check no pending transaction already exists for this expiry period
                existing = Transaction.query.filter(
                    Transaction.member_id == member.id,
                    Transaction.status == TransactionStatus.PENDING,
                    Transaction.due_date == member.package_expiry_date
                ).first()

                if not existing:
                    trainer_fee = 0
                    if member.trainer_id:
                        trainer = TrainerProfile.query.get(member.trainer_id)
                        if trainer and trainer.salary_rate:
                            trainer_fee = float(trainer.salary_rate)

                    package_price = float(package.price) if package.price else 0
                    total_amount = package_price + trainer_fee

                    new_txn = Transaction(
                        member_id=member.id,
                        amount=total_amount,
                        transaction_type=TransactionType.PAYMENT,
                        status=TransactionStatus.PENDING,
                        due_date=member.package_expiry_date,  # due = expiry date
                        package_price=package_price,
                        trainer_fee=trainer_fee,
                        discount_amount=0,
                        discount_type='fixed',
                        created_at=datetime.utcnow()
                    )
                    db.session.add(new_txn)
                    log_action(
                        action='created transaction',
                        target_type='Transaction',
                        target_id=new_txn.id,
                        details={
                            'member_number': member.member_number,
                            'member_name': member.full_name,
                            'amount': total_amount,
                            'transaction_type': 'PAYMENT',
                            'due_date': member.package_expiry_date.isoformat(),
                        }
                    )
        # ─────────────────────────────────────────────────────────────────────────

        db.session.commit()
        
        member_dict = member.to_dict()
        user = User.query.get(member.user_id)
        if user:
            member_dict['username'] = user.username
        
        return jsonify(member_dict), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/members/<member_id>', methods=['DELETE'])
@require_admin
def delete_member(member_id):
    """Delete a member."""
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    try:
        # Log the action before deletion
        log_action(
            action='deleted member',
            target_type='Member',
            target_id=member.id,
            details={
                'member_number': member.member_number,
                'full_name': member.full_name,
                'phone': member.phone,
                'card_id': member.card_id
            }
        )
        
        user = User.query.get(member.user_id)
        db.session.delete(member)
        if user:
            db.session.delete(user)
        db.session.commit()
        
        return jsonify({'message': 'Member deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/members/<member_id>/set-password', methods=['POST'])
@require_admin
def set_member_password(member_id):
    """Set a new password for a member."""
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    data = request.get_json()
    new_password = data.get('password')
    
    if not new_password:
        return jsonify({'error': 'Password is required'}), 400
    
    if len(new_password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters long'}), 400
    
    try:
        user = User.query.get(member.user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Update both hashed password and plain password
        user.password_hash = PasswordService.hash_password(new_password)
        user.password = new_password  # Store plain password for display
        
        db.session.commit()
        
        # Log the action
        log_action(
            action='set member password',
            target_type='Member',
            target_id=member.id,
            details={
                'member_number': member.member_number,
                'full_name': member.full_name
            }
        )
        
        return jsonify({'message': 'Password updated successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/members/<member_id>/reset-password-email', methods=['POST'])
@require_admin
def send_reset_password_email(member_id):
    """Send password reset email to member."""
    member = MemberProfile.query.get(member_id)
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    if not member.email:
        return jsonify({'error': 'Member does not have an email address'}), 400
    
    try:
        user = User.query.get(member.user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Generate reset token
        from services.password_service import PasswordService
        reset_token = PasswordService.generate_reset_token()
        reset_token_expiry = datetime.utcnow() + timedelta(hours=24)
        
        user.reset_token = reset_token
        user.reset_token_expiry = reset_token_expiry
        db.session.commit()
        
        # Send email
        from services.email_service import EmailService
        email_service = EmailService()
        email_service.send_password_reset_email(
            to_email=member.email,
            member_name=member.full_name,
            reset_token=reset_token
        )
        
        # Log the action
        log_action(
            action='sent password reset email',
            target_type='Member',
            target_id=member.id,
            details={
                'member_number': member.member_number,
                'full_name': member.full_name,
                'email': member.email
            }
        )
        
        return jsonify({'message': 'Password reset email sent successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error sending password reset email: {str(e)}")
        return jsonify({'error': 'Failed to send password reset email'}), 500


# ============================================================================
# TRAINER ROUTES
# ============================================================================

@admin_complete_bp.route('/trainers', methods=['GET'])
@require_admin
def list_trainers():
    """List all trainers with pagination."""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    paginated = TrainerProfile.query.paginate(page=page, per_page=per_page, error_out=False)
    
    trainers = []
    for trainer in paginated.items:
        trainer_dict = trainer.to_dict()
        user = User.query.get(trainer.user_id)
        if user:
            trainer_dict['username'] = user.username
        
        # Count assigned members
        assigned_members_count = MemberProfile.query.filter_by(trainer_id=trainer.id).count()
        trainer_dict['assigned_members_count'] = assigned_members_count
        
        trainers.append(trainer_dict)
    
    return jsonify({
        'trainers': trainers,
        'total': paginated.total,
        'page': page,
        'per_page': per_page
    }), 200


@admin_complete_bp.route('/trainers/<trainer_id>', methods=['GET'])
@require_admin
def get_trainer(trainer_id):
    """Get a specific trainer by ID."""
    trainer = TrainerProfile.query.get(trainer_id)
    if not trainer:
        return jsonify({'error': 'Trainer not found'}), 404
    
    trainer_dict = trainer.to_dict()
    user = User.query.get(trainer.user_id)
    if user:
        trainer_dict['username'] = user.username
    
    return jsonify(trainer_dict), 200


@admin_complete_bp.route('/trainers', methods=['POST'])
@require_admin
def create_trainer():
    """Create a new trainer."""
    data = request.get_json()
    
    # Validate required fields (no username/password/email needed)
    required_fields = ['specialization', 'phone']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    try:
        # Create user with auto-generated username (phone-based)
        auto_username = 'trainer_' + data['phone'][-4:] + '_' + str(uuid.uuid4())[:8]
        auto_password = str(uuid.uuid4())
        
        user = User(
            username=auto_username,
            password_hash=PasswordService.hash_password(auto_password),
            role=UserRole.TRAINER,
            is_active=True
        )
        db.session.add(user)
        db.session.flush()
        
        # Parse date_of_birth if provided
        dob = None
        if data.get('date_of_birth'):
            try:
                from datetime import datetime as dt
                dob = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            except:
                dob = None
        
        # Create trainer profile (email is optional, can be None)
        trainer = TrainerProfile(
            user_id=user.id,
            specialization=data['specialization'],
            phone=data['phone'],
            email=data.get('email'),  # Optional
            full_name=data.get('full_name'),
            gender=data.get('gender'),
            date_of_birth=dob,
            cnic=data.get('cnic'),
            salary_rate=data.get('salary_rate', 0),
            availability=data.get('availability')
        )
        db.session.add(trainer)
        
        # Log the action
        log_action(
            action='created trainer',
            target_type='Trainer',
            target_id=trainer.id,
            details={
                'full_name': data.get('full_name'),
                'phone': data['phone'],
                'specialization': data['specialization'],
                'salary_rate': data.get('salary_rate', 0)
            }
        )
        
        db.session.commit()
        
        trainer_dict = trainer.to_dict()
        
        return jsonify(trainer_dict), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/trainers/<trainer_id>', methods=['PUT'])
@require_admin
def update_trainer(trainer_id):
    """Update a trainer."""
    trainer = TrainerProfile.query.get(trainer_id)
    if not trainer:
        return jsonify({'error': 'Trainer not found'}), 404
    
    data = request.get_json()
    
    try:
        # Capture before state
        before_state = {
            'full_name': trainer.full_name,
            'specialization': trainer.specialization,
            'phone': trainer.phone,
            'email': trainer.email,
            'salary_rate': float(trainer.salary_rate) if trainer.salary_rate else 0,
            'availability': trainer.availability
        }
        
        # Update trainer profile fields
        if 'full_name' in data:
            trainer.full_name = data['full_name']
        if 'specialization' in data:
            trainer.specialization = data['specialization']
        if 'phone' in data:
            trainer.phone = data['phone']
        if 'email' in data:
            trainer.email = data['email']
        if 'bio' in data:
            trainer.bio = data['bio']
        if 'gender' in data:
            trainer.gender = data['gender']
        if 'date_of_birth' in data:
            try:
                from datetime import datetime as dt
                trainer.date_of_birth = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date() if data['date_of_birth'] else None
            except:
                pass
        if 'cnic' in data:
            trainer.cnic = data['cnic']
        if 'salary_rate' in data:
            trainer.salary_rate = data['salary_rate']
        if 'availability' in data:
            trainer.availability = data['availability']
        
        # Capture after state
        after_state = {
            'full_name': trainer.full_name,
            'specialization': trainer.specialization,
            'phone': trainer.phone,
            'email': trainer.email,
            'salary_rate': float(trainer.salary_rate) if trainer.salary_rate else 0,
            'availability': trainer.availability
        }
        
        # Log changes
        changes = get_change_details(before_state, after_state)
        if changes:
            log_action(
                action='updated trainer',
                target_type='Trainer',
                target_id=trainer.id,
                details={'changes': changes}
            )
        
        db.session.commit()
        
        trainer_dict = trainer.to_dict()
        user = User.query.get(trainer.user_id)
        if user:
            trainer_dict['username'] = user.username
        
        return jsonify(trainer_dict), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/trainers/<trainer_id>/update-fixed', methods=['PUT'])
@require_admin
def update_trainer_fixed(trainer_id):
    """Update a trainer (alias endpoint for frontend compatibility)."""
    return update_trainer(trainer_id)


@admin_complete_bp.route('/trainers/<trainer_id>', methods=['DELETE'])
@require_admin
def delete_trainer(trainer_id):
    """Delete a trainer and all related records."""
    trainer = TrainerProfile.query.get(trainer_id)
    if not trainer:
        return jsonify({'error': 'Trainer not found'}), 404
    
    try:
        # Log before deletion
        log_action(
            action='deleted trainer',
            target_type='Trainer',
            target_id=trainer.id,
            details={
                'full_name': trainer.full_name,
                'phone': trainer.phone,
                'specialization': trainer.specialization
            }
        )

        # ── Remove trainer from members who have this trainer assigned ──
        MemberProfile.query.filter_by(trainer_id=trainer_id).update(
            {'trainer_id': None}, synchronize_session=False
        )

        # ── Delete trainer commission records (trainer_member_charges) ──
        from models.trainer_commission import TrainerMemberCharge, TrainerSalarySlip
        TrainerMemberCharge.query.filter_by(trainer_id=trainer_id).delete(synchronize_session=False)
        TrainerSalarySlip.query.filter_by(trainer_id=trainer_id).delete(synchronize_session=False)

        # ── Delete the trainer profile and user account ──
        user = User.query.get(trainer.user_id)
        db.session.delete(trainer)
        if user:
            db.session.delete(user)

        db.session.commit()
        
        return jsonify({'message': 'Trainer deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# DASHBOARD METRICS
# ============================================================================

@admin_complete_bp.route('/dashboard/metrics', methods=['GET'])
@require_admin
def get_dashboard_metrics():
    """Get dashboard metrics including inactive members."""
    try:
        from datetime import datetime, timedelta
        from models.attendance import Attendance
        from sqlalchemy import func
        
        # Total active members
        total_members = MemberProfile.query.filter_by(is_frozen=False).count()
        
        # Overdue payments - count PENDING transactions where due_date < now
        # EXCLUDE transactions for frozen members
        now = datetime.utcnow()
        
        # Get all overdue transactions
        overdue_transactions = Transaction.query.filter(
            Transaction.status == TransactionStatus.PENDING,
            Transaction.due_date < now
        ).all()
        
        # Filter out transactions for frozen members
        overdue_payments_count = 0
        for transaction in overdue_transactions:
            member = MemberProfile.query.get(transaction.member_id)
            if member and not member.is_frozen:
                overdue_payments_count += 1
        
        # Inactive members - members who have attended at least once but not in the last 10 days
        # Logic: Only count as inactive if member has attendance history AND no recent attendance
        ten_days_ago = now - timedelta(days=10)
        
        # Get all active (non-frozen) members
        active_members = MemberProfile.query.filter_by(is_frozen=False).all()
        
        # Count members who are inactive (have attended before but not in last 10 days)
        inactive_members_count = 0
        for member in active_members:
            # First check if member has ANY attendance record (ever)
            has_attended_before = Attendance.query.filter(
                Attendance.member_id == member.id
            ).first()
            
            # Only check recent attendance if member has attended at least once
            if has_attended_before:
                # Check if member has any attendance in the last 10 days
                recent_attendance = Attendance.query.filter(
                    Attendance.member_id == member.id,
                    Attendance.check_in_time >= ten_days_ago
                ).first()
                
                # If they've attended before but NOT in last 10 days, they're inactive
                if not recent_attendance:
                    inactive_members_count += 1
        
        # Since attendance was removed from original metrics, set these to 0
        daily_attendance_count = 0
        live_floor_count = 0
        
        result = {
            'total_members': total_members,
            'overdue_payments_count': overdue_payments_count,
            'inactive_members_count': inactive_members_count,
            'daily_attendance_count': daily_attendance_count,
            'live_floor_count': live_floor_count
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/recent-activity', methods=['GET'])
@require_admin
def get_recent_activity():
    """Get recent activity for dashboard - newest members and payments."""
    try:
        activities = []

        # Get 5 most recently created members (sorted by created_at on User)
        recent_members = db.session.query(MemberProfile, User).join(
            User, MemberProfile.user_id == User.id
        ).order_by(User.created_at.desc()).limit(5).all()

        for member, user in recent_members:
            activities.append({
                'type': 'member',
                'icon': 'user',
                'title': 'New member registered',
                'description': f'{member.full_name or user.username} joined the gym',
                'timestamp': user.created_at.isoformat() + 'Z',
            })

        # Get 5 most recently paid transactions
        recent_payments = Transaction.query.filter(
            Transaction.status == TransactionStatus.COMPLETED,
            Transaction.paid_date.isnot(None)
        ).order_by(Transaction.paid_date.desc()).limit(5).all()

        for txn in recent_payments:
            activities.append({
                'type': 'payment',
                'icon': 'payment',
                'title': 'Payment received',
                'description': f'Rs. {int(txn.amount):,} - Membership fee',
                'timestamp': txn.paid_date.isoformat() + 'Z',
            })

        # Sort combined list by timestamp newest first, return top 10
        activities.sort(key=lambda x: x['timestamp'], reverse=True)

        return jsonify({'activities': activities[:10]}), 200

    except Exception as e:
        current_app.logger.error(f"Error fetching recent activity: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ============================================================================
# SETTINGS ROUTES
# ============================================================================

@admin_complete_bp.route('/settings', methods=['GET'])
@require_admin
def get_admin_settings():
    """Get system settings."""
    from models.settings import Settings
    
    try:
        # Get system settings (should be only one record)
        settings = Settings.query.first()
        if not settings:
            # Create default settings if none exist
            settings = Settings(admission_fee=5000)
            db.session.add(settings)
            db.session.commit()
        
        return jsonify({
            'admission_fee': float(settings.admission_fee),
            'package_price': 0,
            'package_duration': 30,
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/settings', methods=['PUT'])
@require_admin
def update_admin_settings():
    """Update system settings."""
    from models.settings import Settings
    
    data = request.get_json()
    
    try:
        # Get or create settings record
        settings = Settings.query.first()
        if not settings:
            settings = Settings()
            db.session.add(settings)
        
        # Update admission_fee if provided
        if 'admission_fee' in data:
            settings.admission_fee = float(data['admission_fee'])
        
        db.session.commit()
        
        return jsonify({
            'message': 'Settings updated successfully',
            'admission_fee': float(settings.admission_fee),
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/profile', methods=['PUT'])
@require_admin
def update_admin_profile():
    """Update admin profile."""
    from flask_jwt_extended import get_jwt_identity
    
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    data = request.get_json()
    
    try:
        # Update password if provided
        if 'current_password' in data and 'new_password' in data:
            if not PasswordService.verify_password(data['current_password'], user.password_hash):
                return jsonify({'error': 'Current password is incorrect'}), 400
            
            user.password_hash = PasswordService.hash_password(data['new_password'])
        
        # Update username if provided
        if 'username' in data and data['username'] != user.username:
            existing = User.query.filter_by(username=data['username']).first()
            if existing:
                return jsonify({'error': 'Username already exists'}), 400
            user.username = data['username']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Profile updated successfully',
            'user': user.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500



# ============================================================================
# ANALYTICS ENDPOINTS
# ============================================================================

@admin_complete_bp.route('/dashboard/revenue-projection', methods=['GET'])
@require_admin
def get_revenue_projection():
    """Get revenue projection data based on active member packages."""
    try:
        # Get all active members (not frozen) with packages
        active_members = MemberProfile.query.filter(
            MemberProfile.is_frozen == False,
            MemberProfile.current_package_id.isnot(None)
        ).all()
        
        projected_monthly_revenue = 0
        active_packages_count = 0
        now = datetime.now()
        
        for member in active_members:
            # Check if package is still active (not expired or no expiry date set)
            if member.package_expiry_date is None or member.package_expiry_date >= now:
                package = Package.query.get(member.current_package_id)
                if package and package.is_active:
                    # Calculate monthly revenue based on package duration
                    if package.duration_days > 0:
                        monthly_rate = (float(package.price) / package.duration_days) * 30
                        projected_monthly_revenue += monthly_rate
                        active_packages_count += 1
        
        return jsonify({
            'projected_monthly_revenue': round(projected_monthly_revenue, 2),
            'active_packages_count': active_packages_count,
            'current_revenue': round(projected_monthly_revenue, 2),
            'growth_rate': 0
        }), 200
    except Exception as e:
        print(f"Error in get_revenue_projection: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/revenue-trend', methods=['GET'])
@require_admin
def get_revenue_trend():
    """Get revenue trend data."""
    try:
        from datetime import datetime, timedelta
        
        # Get last 6 months of revenue
        months = []
        revenue = []
        
        for i in range(6):
            month_date = datetime.now() - timedelta(days=30*i)
            months.insert(0, month_date.strftime('%b'))
            
            # Calculate revenue for this month
            month_start = month_date.replace(day=1)
            if i > 0:
                next_month = (month_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            else:
                next_month = datetime.now()
            
            month_revenue = db.session.query(func.sum(Transaction.amount)).filter(
                Transaction.status == TransactionStatus.COMPLETED,
                Transaction.paid_date >= month_start,
                Transaction.paid_date < next_month
            ).scalar() or 0
            
            revenue.insert(0, float(month_revenue))
        
        return jsonify({
            'months': months,
            'revenue': revenue
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/member-growth', methods=['GET'])
@require_admin
def get_member_growth():
    """Get member growth data."""
    try:
        from datetime import datetime, timedelta
        
        # Get last 6 months of member growth
        months = []
        counts = []
        
        for i in range(6):
            month_date = datetime.now() - timedelta(days=30*i)
            months.insert(0, month_date.strftime('%b'))
            
            # Count members created up to this month
            month_end = month_date.replace(day=28) + timedelta(days=4)
            month_end = month_end.replace(day=1)
            
            count = User.query.filter(
                User.role == UserRole.MEMBER,
                User.created_at < month_end
            ).count()
            
            counts.insert(0, count)
        
        return jsonify({
            'months': months,
            'counts': counts
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/attendance-trend', methods=['GET'])
@require_admin
def get_attendance_trend():
    """Get attendance trend data (returns empty since attendance was removed)."""
    try:
        from datetime import datetime, timedelta
        
        # Return empty data since attendance was removed
        months = []
        attendance_count = []
        
        for i in range(6):
            month_date = datetime.now() - timedelta(days=30*i)
            months.insert(0, month_date.strftime('%b'))
            attendance_count.insert(0, 0)
        
        return jsonify({
            'months': months,
            'attendance_count': attendance_count
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/dashboard/peak-hours', methods=['GET'])
@require_admin
def get_peak_hours():
    """Get peak hours data (returns empty since attendance was removed)."""
    try:
        # Return empty data since attendance was removed
        peak_hours = []
        
        for hour in range(6, 23):  # 6 AM to 10 PM
            peak_hours.append({
                'hour': hour,
                'count': 0
            })
        
        return jsonify({
            'peak_hours': peak_hours
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# FINANCE ENDPOINTS (for AdminFinance page) - OPTIMIZED
# ============================================================================

@admin_complete_bp.route('/finance/overdue-members', methods=['GET'])
@require_admin
def get_overdue_members():
    """Get members with overdue payments AND grace period members - optimized for Finance dashboard."""
    try:
        from models import Settings
        from sqlalchemy.orm import joinedload
        
        # Get grace period from settings
        settings = Settings.query.first()
        grace_period_days = settings.grace_period_days if settings else 3
        
        # Calculate grace period cutoff date (truly overdue)
        grace_cutoff = datetime.utcnow() - timedelta(days=grace_period_days)
        
        # Get ALL pending transactions that are past due date (includes grace period + overdue)
        all_pending_transactions = Transaction.query.options(
            joinedload(Transaction.member)
        ).filter(
            Transaction.status == TransactionStatus.PENDING,
            Transaction.due_date < datetime.utcnow()
        ).order_by(Transaction.due_date.asc()).all()
        
        # Group by member and categorize
        members_dict = {}
        for transaction in all_pending_transactions:
            member = transaction.member
            if not member:
                continue
            
            member_id = member.id
            if member_id not in members_dict:
                # Calculate total amount and oldest due date for this member
                member_transactions = [t for t in all_pending_transactions if t.member_id == member_id]
                total_amount = sum(t.amount for t in member_transactions)
                oldest_due_date = min(t.due_date for t in member_transactions)
                
                # Calculate days since due date
                days_since_due = (datetime.utcnow() - oldest_due_date).days
                
                # Determine status: grace period or overdue
                if days_since_due <= grace_period_days:
                    # In grace period
                    status = 'grace'
                    grace_day = days_since_due
                    days_overdue = 0
                else:
                    # Truly overdue (past grace period)
                    status = 'overdue'
                    grace_day = 0
                    days_overdue = days_since_due - grace_period_days
                
                members_dict[member_id] = {
                    'member_id': member.id,
                    'member_number': member.member_number,
                    'full_name': member.full_name,
                    'phone': member.phone,
                    'profile_picture': member.profile_picture,
                    'total_overdue': float(total_amount),
                    'overdue_count': len(member_transactions),
                    'days_overdue': days_overdue,
                    'status': status,
                    'grace_day': grace_day,
                    'grace_period_total': grace_period_days,
                    'oldest_due_date': oldest_due_date.isoformat() + 'Z',
                    'transactions': [t.id for t in member_transactions]
                }
        
        all_members = list(members_dict.values())
        
        # Separate into grace and overdue
        grace_members = [m for m in all_members if m['status'] == 'grace']
        overdue_members = [m for m in all_members if m['status'] == 'overdue']
        
        return jsonify({
            'overdue_members': overdue_members,
            'grace_members': grace_members,
            'all_members': all_members,
            'total_count': len(all_members),
            'overdue_count': len(overdue_members),
            'grace_count': len(grace_members),
            'total_amount': sum(m['total_overdue'] for m in all_members)
        }), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching overdue members: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/monthly-summary', methods=['GET'])
@require_admin
def get_monthly_summary():
    """Get financial summary grouped by month - last 3 months by default."""
    try:
        months = request.args.get('months', 3, type=int)
        
        # Calculate date range
        today = datetime.utcnow()
        start_date = today.replace(day=1) - timedelta(days=30 * (months - 1))
        start_date = start_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Get all transactions in date range
        transactions = Transaction.query.filter(
            Transaction.due_date >= start_date
        ).all()
        
        # Group by month
        monthly_data = {}
        for transaction in transactions:
            if not transaction.due_date:
                continue
            
            month_key = transaction.due_date.strftime('%Y-%m')
            month_name = transaction.due_date.strftime('%B %Y')
            
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    'month': month_key,
                    'month_name': month_name,
                    'total_collected': 0,
                    'total_pending': 0,
                    'total_overdue': 0,
                    'completed_count': 0,
                    'pending_count': 0,
                    'overdue_count': 0
                }
            
            amount = float(transaction.amount)
            
            if transaction.status == TransactionStatus.COMPLETED:
                monthly_data[month_key]['total_collected'] += amount
                monthly_data[month_key]['completed_count'] += 1
            elif transaction.status == TransactionStatus.PENDING:
                # Check if overdue
                if transaction.due_date < datetime.utcnow():
                    monthly_data[month_key]['total_overdue'] += amount
                    monthly_data[month_key]['overdue_count'] += 1
                else:
                    monthly_data[month_key]['total_pending'] += amount
                    monthly_data[month_key]['pending_count'] += 1
        
        # Convert to sorted list (most recent first)
        summary = sorted(monthly_data.values(), key=lambda x: x['month'], reverse=True)
        
        return jsonify({'monthly_summary': summary}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching monthly summary: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/transactions-by-month', methods=['GET'])
@require_admin
def get_transactions_by_month():
    """Get transactions for a specific month - optimized."""
    try:
        month = request.args.get('month')  # Format: YYYY-MM
        
        if not month:
            return jsonify({'error': 'Month parameter required (format: YYYY-MM)'}), 400
        
        # Parse month
        try:
            year, month_num = map(int, month.split('-'))
            start_date = datetime(year, month_num, 1)
            
            # Calculate end date (first day of next month)
            if month_num == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month_num + 1, 1)
        except ValueError:
            return jsonify({'error': 'Invalid month format. Use YYYY-MM'}), 400
        
        # Get transactions for the month with member data (eager loading)
        transactions = Transaction.query.filter(
            Transaction.due_date >= start_date,
            Transaction.due_date < end_date
        ).order_by(Transaction.due_date.desc()).all()
        
        # Format transactions with member details
        payments = []
        for transaction in transactions:
            member = MemberProfile.query.get(transaction.member_id)
            if not member:
                continue
            
            # Calculate status
            status = transaction.status.value
            if status != 'COMPLETED' and transaction.due_date:
                if datetime.utcnow() > transaction.due_date:
                    status = 'OVERDUE'
            
            payments.append({
                'id': transaction.id,
                'member_id': transaction.member_id,
                'member_number': member.member_number,
                'full_name': member.full_name,
                'phone': member.phone,
                'amount': float(transaction.amount),
                'transaction_type': transaction.transaction_type.value if hasattr(transaction.transaction_type, 'value') else str(transaction.transaction_type),
                'status': status,
                'due_date': transaction.due_date.isoformat() + 'Z' if transaction.due_date else None,
                'paid_date': transaction.paid_date.isoformat() + 'Z' if transaction.paid_date else None,
                'created_at': transaction.created_at.isoformat() + 'Z'
            })
        
        return jsonify({
            'transactions': payments,
            'month': month,
            'total_count': len(payments)
        }), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching transactions by month: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/export-transactions', methods=['GET'])
@require_admin
def export_transactions_excel():
    """Export transactions for a specific month to Excel file with GOFIT theme."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from sqlalchemy.orm import joinedload
        
        month = request.args.get('month')  # Format: YYYY-MM
        
        if not month:
            return jsonify({'error': 'Month parameter required (format: YYYY-MM)'}), 400
        
        # Parse month
        try:
            year, month_num = map(int, month.split('-'))
            start_date = datetime(year, month_num, 1)
            
            # Calculate end date (first day of next month)
            if month_num == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month_num + 1, 1)
        except ValueError:
            return jsonify({'error': 'Invalid month format. Use YYYY-MM'}), 400
        
        # Get transactions for the month with member data (eager loading)
        transactions = Transaction.query.options(
            joinedload(Transaction.member)
        ).filter(
            Transaction.due_date >= start_date,
            Transaction.due_date < end_date
        ).order_by(Transaction.due_date.desc()).all()
        
        if not transactions:
            return jsonify({'error': 'No transactions found for this month'}), 404
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Finance Report"
        
        # Define headers
        headers = [
            'Member ID', 'Member Name', 'Phone', 'Type', 
            'Amount (Rs.)', 'Status', 'Due Date', 'Paid Date'
        ]
        
        # GOFIT theme colors
        header_fill = PatternFill(start_color='F2C228', end_color='F2C228', fill_type='solid')
        header_font = Font(bold=True, size=12, color='1A1A1A')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Reusable styles
        data_alignment_left = Alignment(horizontal='left', vertical='center')
        data_alignment_right = Alignment(horizontal='right', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        data_font = Font(size=11, color='1A1A1A')
        border = Border(
            left=Side(style='thin', color='1A1A1A'),
            right=Side(style='thin', color='1A1A1A'),
            top=Side(style='thin', color='1A1A1A'),
            bottom=Side(style='thin', color='1A1A1A')
        )
        
        # Status colors
        completed_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        pending_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        overdue_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        row_fill_even = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        # Add title row
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        month_date = datetime(year, month_num, 1)
        month_text = month_date.strftime('%B %Y')
        title_cell.value = f'GOFIT - FINANCE REPORT - {month_text}'
        title_cell.font = Font(bold=True, size=16, color='1A1A1A')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = header_fill
        ws.row_dimensions[1].height = 30
        
        # Add headers in row 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        ws.row_dimensions[2].height = 25
        
        # Add data rows
        total_collected = 0
        total_pending = 0
        total_overdue = 0
        completed_count = 0
        pending_count = 0
        overdue_count = 0
        
        for row_num, transaction in enumerate(transactions, 3):
            member = transaction.member
            if not member:
                continue
            
            # Calculate status
            status = transaction.status.value
            if status != 'COMPLETED' and transaction.due_date:
                if datetime.utcnow() > transaction.due_date:
                    status = 'OVERDUE'
            
            # Format dates
            due_date = transaction.due_date.strftime('%d-%b-%Y') if transaction.due_date else 'N/A'
            paid_date = transaction.paid_date.strftime('%d-%b-%Y') if transaction.paid_date else 'N/A'
            
            # Transaction type
            txn_type = transaction.transaction_type.value if hasattr(transaction.transaction_type, 'value') else str(transaction.transaction_type)
            
            # Row data
            row_data = [
                member.member_number or 'N/A',
                member.full_name or 'N/A',
                member.phone or 'N/A',
                txn_type,
                float(transaction.amount),
                status,
                due_date,
                paid_date
            ]
            
            # Determine row fill based on status
            if status == 'COMPLETED':
                row_fill = completed_fill
                total_collected += float(transaction.amount)
                completed_count += 1
            elif status == 'OVERDUE':
                row_fill = overdue_fill
                total_overdue += float(transaction.amount)
                overdue_count += 1
            else:
                row_fill = pending_fill
                total_pending += float(transaction.amount)
                pending_count += 1
            
            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.fill = row_fill
                cell.font = data_font
                
                # Alignment based on column
                if col_num == 5:  # Amount
                    cell.alignment = data_alignment_right
                    cell.number_format = '#,##0.00'
                elif col_num in [6, 7, 8]:  # Status, dates
                    cell.alignment = data_alignment_center
                else:
                    cell.alignment = data_alignment_left
        
        # Add summary section
        summary_row = len(transactions) + 3
        
        # Summary title
        ws.merge_cells(f'A{summary_row}:H{summary_row}')
        summary_title = ws[f'A{summary_row}']
        summary_title.value = 'SUMMARY'
        summary_title.font = Font(bold=True, size=14, color='1A1A1A')
        summary_title.alignment = Alignment(horizontal='center', vertical='center')
        summary_title.fill = header_fill
        summary_title.border = border
        ws.row_dimensions[summary_row].height = 25
        
        # Summary rows
        summary_data = [
            ('Total Collected', total_collected, completed_count, completed_fill),
            ('Total Pending', total_pending, pending_count, pending_fill),
            ('Total Overdue', total_overdue, overdue_count, overdue_fill),
            ('GRAND TOTAL', total_collected + total_pending + total_overdue, len(transactions), header_fill)
        ]
        
        for i, (label, amount, count, fill) in enumerate(summary_data, 1):
            row = summary_row + i
            
            # Label
            ws.merge_cells(f'A{row}:D{row}')
            label_cell = ws[f'A{row}']
            label_cell.value = label
            label_cell.font = Font(bold=True, size=12, color='1A1A1A')
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            label_cell.fill = fill
            label_cell.border = border
            
            # Amount
            ws.merge_cells(f'E{row}:F{row}')
            amount_cell = ws[f'E{row}']
            amount_cell.value = amount
            amount_cell.font = Font(bold=True, size=12, color='1A1A1A')
            amount_cell.alignment = data_alignment_right
            amount_cell.fill = fill
            amount_cell.border = border
            amount_cell.number_format = '#,##0.00'
            
            # Count
            ws.merge_cells(f'G{row}:H{row}')
            count_cell = ws[f'G{row}']
            count_cell.value = f'{count} transactions'
            count_cell.font = Font(bold=True, size=11, color='1A1A1A')
            count_cell.alignment = data_alignment_center
            count_cell.fill = fill
            count_cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"GOFIT_Finance_{month}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        current_app.logger.error(f"Error exporting transactions to Excel: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/member-payments-fixed', methods=['GET'])
@require_admin
def get_member_payments_fixed():
    """Get all transactions with proper status calculation."""
    try:
        # Get all transactions
        transactions = Transaction.query.all()
        
        payments = []
        for transaction in transactions:
            try:
                member = MemberProfile.query.get(transaction.member_id)
                if not member:
                    continue
                
                user = User.query.get(member.user_id)
                username = user.username if user else 'Unknown'
                
                # Calculate status based on transaction state and dates
                status = transaction.status.value
                
                # If not completed, check if overdue
                if status != 'COMPLETED' and transaction.due_date:
                    if datetime.utcnow() > transaction.due_date:
                        status = 'OVERDUE'
                
                # Safely get trainer_fee and discount_amount
                trainer_fee = 0
                discount_amount = 0
                package_price = 0
                try:
                    trainer_fee = float(transaction.trainer_fee) if transaction.trainer_fee else 0
                except (TypeError, ValueError):
                    trainer_fee = 0
                
                try:
                    discount_amount = float(transaction.discount_amount) if transaction.discount_amount else 0
                except (TypeError, ValueError):
                    discount_amount = 0
                
                try:
                    package_price = float(transaction.package_price) if transaction.package_price else 0
                except (TypeError, ValueError):
                    package_price = 0
                
                payments.append({
                    'id': transaction.id,
                    'member_id': transaction.member_id,
                    'username': username,
                    'full_name': member.full_name,
                    'phone': member.phone,
                    'amount': float(transaction.amount),
                    'transaction_type': transaction.transaction_type.value if hasattr(transaction.transaction_type, 'value') else str(transaction.transaction_type),
                    'status': status,
                    'due_date': transaction.due_date.isoformat() + 'Z' if transaction.due_date else None,
                    'paid_date': transaction.paid_date.isoformat() + 'Z' if transaction.paid_date else None,
                    'created_at': transaction.created_at.isoformat() + 'Z',
                    'total_monthly_payment': float(transaction.amount),
                    'trainer_fee': trainer_fee,
                    'package_price': package_price,
                    'discount_amount': discount_amount,
                    'discount_type': transaction.discount_type or 'fixed'
                })
            except Exception as item_error:
                continue
        
        return jsonify({'payments': payments}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/transactions/<transaction_id>/mark-paid', methods=['POST'])
@require_admin
def mark_transaction_paid(transaction_id):
    """Mark a transaction as paid."""
    try:
        transaction = Transaction.query.get(transaction_id)
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404
        
        # Check if transaction is already completed - prevent duplicate processing
        if transaction.status == TransactionStatus.COMPLETED:
            return jsonify({
                'message': 'Transaction is already marked as paid',
                'transaction': transaction.to_dict()
            }), 200
        
        # Get member info
        member = MemberProfile.query.get(transaction.member_id)
        if not member:
            return jsonify({'error': 'Member not found'}), 404
        
        # Update transaction status and paid date
        transaction.status = TransactionStatus.COMPLETED
        transaction.paid_date = datetime.utcnow()
        
        # If this is an admission fee, mark admission_fee_paid on member
        if transaction.transaction_type.value == 'ADMISSION':
            if member:
                member.admission_fee_paid = True
        
        # Log the action
        log_action(
            action='marked transaction paid',
            target_type='Transaction',
            target_id=transaction.id,
            details={
                'member_number': member.member_number if member else None,
                'member_name': member.full_name if member else None,
                'amount': float(transaction.amount),
                'transaction_type': transaction.transaction_type.value,
                'paid_date': transaction.paid_date.isoformat() if transaction.paid_date else None
            }
        )
        
        db.session.commit()
        
        return jsonify({
            'message': 'Transaction marked as paid successfully',
            'transaction': transaction.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_complete_bp.route('/finance/members/<member_id>/create-transaction', methods=['POST'])
@require_admin
def create_member_transaction(member_id):
    """Create a new transaction for a member (monthly payment)."""
    try:
        member = MemberProfile.query.get(member_id)
        if not member:
            return jsonify({'error': 'Member not found'}), 404
        
        # Check if member has an active package
        if not member.current_package_id:
            return jsonify({'error': 'Member does not have an active package'}), 400
        
        package = Package.query.get(member.current_package_id)
        if not package or not package.is_active:
            return jsonify({'error': 'Member package is not active'}), 400
        
        # Get the last transaction to calculate next due date
        last_transaction = Transaction.query.filter_by(
            member_id=member_id
        ).order_by(Transaction.due_date.desc()).first()
        
        # Calculate next due date (30 days from last transaction or today)
        if last_transaction and last_transaction.due_date:
            next_due_date = last_transaction.due_date + timedelta(days=30)
        else:
            next_due_date = datetime.utcnow() + timedelta(days=30)
        
        # Check if a transaction already exists for this due date
        date_range_start = next_due_date - timedelta(days=5)
        date_range_end = next_due_date + timedelta(days=5)
        
        existing_transaction = Transaction.query.filter(
            Transaction.member_id == member_id,
            Transaction.due_date >= date_range_start,
            Transaction.due_date <= date_range_end
        ).first()
        
        if existing_transaction:
            return jsonify({'error': 'A transaction already exists for this period'}), 400
        
        # Get trainer fee if trainer is assigned
        trainer_fee = 0
        if member.trainer_id:
            trainer = TrainerProfile.query.get(member.trainer_id)
            if trainer:
                trainer_fee = float(trainer.salary_rate) if trainer.salary_rate else 0
        
        # Calculate total amount
        package_price = float(package.price) if package.price else 0
        total_amount = package_price + trainer_fee
        
        # Create new transaction
        new_transaction = Transaction(
            member_id=member_id,
            amount=total_amount,
            transaction_type=TransactionType.PAYMENT,
            status=TransactionStatus.PENDING,
            due_date=next_due_date,
            trainer_fee=trainer_fee,
            package_price=package_price,
            discount_amount=0,
            discount_type='fixed',
            created_at=datetime.utcnow()
        )
        
        db.session.add(new_transaction)
        
        # Log the action
        log_action(
            action='created transaction',
            target_type='Transaction',
            target_id=new_transaction.id,
            details={
                'member_number': member.member_number,
                'member_name': member.full_name,
                'amount': total_amount,
                'transaction_type': 'PAYMENT',
                'due_date': next_due_date.isoformat(),
                'package_price': package_price,
                'trainer_fee': trainer_fee
            }
        )
        
        db.session.commit()
        
        return jsonify({
            'message': 'Transaction created successfully',
            'transaction': new_transaction.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# TRAINERS ENDPOINTS (for AdminTrainers page)
# ============================================================================

@admin_complete_bp.route('/trainers-fixed', methods=['GET'])
@require_admin
def list_trainers_fixed():
    """List all trainers (alias for /trainers endpoint)."""
    return list_trainers()


# ============================================================================
# ADMIN SETTINGS ENDPOINTS
# ============================================================================

@admin_complete_bp.route('/change-password', methods=['POST'])
@jwt_required()
@require_admin
def change_password():
    """Change admin password."""
    try:
        from flask_jwt_extended import get_jwt_identity
        
        data = request.get_json()
        
        if not data or not data.get('current_password') or not data.get('new_password'):
            return jsonify({'error': 'Missing required fields'}), 400
        
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        # Validate new password length
        if len(new_password) < 6:
            return jsonify({'error': 'Password must be at least 6 characters'}), 400
        
        # Get current user
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Verify current password
        if not PasswordService.verify_password(current_password, user.password_hash):
            return jsonify({'error': 'Current password is incorrect'}), 401
        
        # Update password
        user.password_hash = PasswordService.hash_password(new_password)
        db.session.commit()
        
        return jsonify({'message': 'Password changed successfully'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"Error in change_password: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
